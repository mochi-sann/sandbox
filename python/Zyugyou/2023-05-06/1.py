
import openpyxl

# 新しいワークブックを作成する
workbook = openpyxl.Workbook()

# シートを取得する
sheet = workbook.active
size = 100

# 行と列のラベルを設定する
sheet['A1'] = ''

for i in range(1, size):
    sheet.cell(row=1, column=i + 1, value=i)
    sheet.cell(row=i + 1, column=1, value=i)

# 掛け算の表を計算する
for i in range(1, size):
    row = i + 1
    for j in range(1, size):
        col = j + 1

        # セルに掛け算の結果を設定する
        sheet.cell(row=row, column=col, value=i * j)

# ファイルを保存する
workbook.save('multiplication_table.xlsx')
